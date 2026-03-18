"""
app/services/export_service.py — Exportación de cartera a XLSX y CSV
"""

import csv
from datetime import datetime
from pathlib import Path

COLUMNAS_EXPORT = [
    ("cliente",            "Cliente"),
    ("factura_no",         "Factura N°"),
    ("fecha_emision",      "Fecha Emisión"),
    ("fecha_vencimiento",  "Fecha Vencimiento"),
    ("monto",              "Total"),
    ("monto_pendiente",    "Saldo Pendiente"),
    ("tipo",               "Estado"),
    ("email",              "Email"),
    ("telefono",           "Teléfono"),
]


def exportar_xlsx(facturas: list[dict], ruta_destino: str | Path) -> tuple[bool, str]:
    """Exporta a XLSX con formato. Retorna (ok, mensaje_error)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return False, "Instala openpyxl: pip install openpyxl>=3.1.0"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cartera"

    header_fill = PatternFill("solid", fgColor="286983")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col_idx, (_, titulo) in enumerate(COLUMNAS_EXPORT, 1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, f in enumerate(facturas, 2):
        for col_idx, (campo, _) in enumerate(COLUMNAS_EXPORT, 1):
            val = f.get(campo, "")
            if campo in ("monto", "monto_pendiente"):
                try:
                    val = round(float(val), 2)
                except (ValueError, TypeError):
                    val = 0.0
            ws.cell(row=row_idx, column=col_idx, value=val)

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

    meta_row = len(facturas) + 3
    ws.cell(row=meta_row, column=1,
            value=f"Exportado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  —  {len(facturas)} registros")
    ws.cell(row=meta_row, column=1).font = Font(color="9893A5", italic=True, size=10)

    try:
        wb.save(str(ruta_destino))
        return True, ""
    except Exception as e:
        return False, str(e)


def exportar_csv(facturas: list[dict], ruta_destino: str | Path) -> tuple[bool, str]:
    """Exporta a CSV UTF-8 con BOM (compatible con Excel en Windows)."""
    try:
        titulos = [titulo for _, titulo in COLUMNAS_EXPORT]
        campos  = [campo  for campo, _ in COLUMNAS_EXPORT]
        with open(ruta_destino, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(titulos)
            for factura in facturas:
                writer.writerow([factura.get(c, "") for c in campos])
        return True, ""
    except Exception as e:
        return False, str(e)
